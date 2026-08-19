#!/usr/bin/env python3
"""
score.py — the authoritative scorer for the ApexChem route-agent exercise.

This file is the scorer. Do not substitute your own — a number it did not produce
does not mean what you think it means.

Usage
-----
    python3 score.py expected.jsonl actual.jsonl [--json]
    python3 score.py --validate actual.jsonl schema.json
    python3 score.py --validate-key expected.jsonl
    python3 score.py --results expected.jsonl actual.jsonl > results.json
    python3 score.py --brier forecasts.json results.json
    python3 score.py --agreement expected.jsonl run1.jsonl run2.jsonl run3.jsonl
    python3 score.py --audit actual.jsonl corpus.xlsx [more.xlsx ...]   # pass BOTH workbooks
    python3 score.py --applies design_requests.jsonl actual.jsonl

Requires Python 3.7+ and, for two modes only, two packages:
    pip install jsonschema     # --validate
    pip install openpyxl       # --audit
Every other mode is pure standard library.

expected.jsonl — one object per line. The shape, with a made-up case:
    {"request_id": "EXAMPLE-00",
     "verdict": "infeasible",              # or a list of acceptable verdicts
     "conflicts": [{"kind": "order_of_operations",
                    "min_severity": "major",       # floor; default "major"
                    "resolution_required": true}], # default: false if floor is blocking
     "forbidden_kinds": ["site_invalid"],  # kinds that are positively WRONG here
     "alternates": [{"verdict": "feasible_with_changes"}],
     "resolved_sequence": "ACDEFGH",       # optional, normalised exact match
     "site_map": [["K5", "K5"]],           # optional, pairs that must appear
     "clean": false, "negative_control": false}
A kind may not appear in both `conflicts` and `forbidden_kinds`; --validate-key
rejects a key that does that.

Scoring, per case, 0 / 1 / 2
    2  verdict acceptable AND every expected conflict kind present at or above
       its floor severity AND every required kind carries a real resolution AND
       no forbidden kind raised AND the answer was discriminate (see below)
    1  one of those halves but not both
    0  otherwise
    A verdict/severity contradiction caps the case at 1.
    A missing or malformed output scores 0 and counts against every metric.

    A correct verdict on its own pays 1 only where the key expects no conflict,
    or where at least one expected conflict was also found. Naming the verdict
    with none of the reasoning behind it is a guess, and a constant guess is
    not an answer.

    A `negative_control` case is passed by REFUSING. Any non-refusal verdict is
    capped at 1 (`+neg_control_not_refused`) even when every clause above holds,
    because no scorer can check that a real alternative route was named. Those
    cases go to the hand-read.

**Raising a conflict the key does not list does not cost points — up to a point.**
Seeing more than the key is the thing this exercise is trying to detect, so
extras are reported as `unexpected_kinds` for a human to read and are never
charged as false positives. Two limits keep that from collapsing into "raise
everything":

  * *Coherence.* A `blocking` finding alongside a non-refusal verdict, or any
    major/blocking finding alongside `feasible`, caps the case at 1. That cap is
    about the verdict contradicting the finding. Raise extras at `minor`, or move
    the verdict.
  * *Discrimination.* Raising more than **two kinds beyond what the key demands**
    forfeits the 2/2 for that case (`+indiscriminate` in the reason). Two is
    generous room to be right about something the key missed; six is a switch
    left on.

A `resolution`, `description` or `detail` string that the submission reuses on
three or more cases is a template, not a resolution, and does not satisfy `resolution_required`. The
comparison is made after NFKC normalisation and stripping to alphanumerics, so
invisible padding does not disguise one template as nine findings.
`forbidden_kinds` — findings that are positively wrong here — cost points at any
severity.

`shotgun_index` is the mean number of unexpected kinds per case. A high value
alongside a low `clean_case_precision` is the signature of an agent flagging
everything. `unreadable_conflict_kinds` lists conflicts dropped because their
`kind` was not one of the enum values: a finding the scorer cannot read is
reported, never silently forgiven.
"""
import json
import re
import signal
import sys
import unicodedata
from collections import Counter
from pathlib import Path

VERDICTS = {"feasible", "feasible_with_changes", "infeasible", "insufficient_information"}
KINDS = {"protecting_group_orthogonality", "order_of_operations", "mutually_exclusive",
         "site_invalid", "reagent_incompatibility", "building_block_availability",
         "intent_not_achieved"}
SEVERITIES = {"blocking", "major", "minor"}
RANK = {"minor": 0, "major": 1, "blocking": 2}
REFUSALS = {"infeasible", "insufficient_information"}


def die(msg):
    print(f"score.py: {msg}", file=sys.stderr)
    sys.exit(2)


def as_list(v):
    """A JSON value that should be a list, coerced defensively."""
    return v if isinstance(v, list) else ([] if v is None else [v])


def dicts(v):
    return [x for x in as_list(v) if isinstance(x, dict)]


def hashable(v):
    return isinstance(v, (str, int, float, bool)) or v is None


def norm(s):
    if s is None:
        return None
    return re.sub(r"\s*([,-])\s*", r"\1", str(s).strip().upper())


def sev_of(c):
    s = c.get("severity")
    if isinstance(s, str):
        s = s.strip().lower()                     # "Major" and "major" are the same claim
    return s if hashable(s) and s in SEVERITIES else "major"   # unknown or absent -> major


def kind_of(c):
    k = c.get("kind")
    return k.strip().lower() if isinstance(k, str) else k


def clean_conflicts(act):
    if not isinstance(act, dict):
        return []
    return [{**c, "kind": kind_of(c)} for c in dicts(act.get("conflicts"))
            if hashable(kind_of(c)) and kind_of(c) in KINDS]


def unknown_kinds(act):
    """Conflicts thrown away because their kind is not one of KINDS.
    Reported, never silently dropped: an unreadable finding must not become a free pass."""
    if not isinstance(act, dict):
        return []
    return sorted({repr(c.get("kind")) for c in dicts(act.get("conflicts"))
                   if not (hashable(kind_of(c)) and kind_of(c) in KINDS)})


def flatten_text(v):
    """Normalise a free-text field for reuse detection: NFKC, case-folded, and stripped
    down to alphanumerics, so zero-width padding and punctuation jitter do not disguise
    one template as nine findings."""
    if not isinstance(v, str):
        return None
    t = unicodedata.normalize("NFKC", v).lower()
    t = re.sub(r"[^a-z0-9]+", " ", t).strip()
    return t or None


def boilerplate(actual, n_cases):
    """Strings a submission reuses across cases. A resolution or a detail that recurs
    on three or more cases is a template, not a finding, and does not satisfy
    `resolution_required`."""
    seen = {}
    for a in actual:
        if not isinstance(a, dict):
            continue
        here = set()
        for c in clean_conflicts(a):
            for fld in ("resolution", "detail", "description"):
                t = flatten_text(c.get(fld))
                if t:
                    here.add(t)
        for t in here:
            seen[t] = seen.get(t, 0) + 1
    cut = 3
    return frozenset(t for t, k in seen.items() if k >= cut)


def accepted(exp):
    v = exp.get("verdict")
    vs = v if isinstance(v, list) else [v]
    return {x if hashable(x) else json.dumps(x, sort_keys=True, default=str) for x in vs}


def load(path, kind):
    out = []
    try:
        with open(path, encoding="utf-8-sig") as fh:   # tolerate a BOM from Windows editors
            lines = list(enumerate(fh, 1))
    except UnicodeDecodeError as e:
        die(f"{path}: not valid UTF-8 ({e})")
    except OSError as e:
        die(str(e))
    for n, line in lines:
        if not line.strip():
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError as e:
            if kind == "expected":
                die(f"{path} line {n}: {e}")
            print(f"score.py: {path} line {n} unparseable, dropped", file=sys.stderr)
            continue
        if not isinstance(row, dict):
            msg = f"{path} line {n}: top-level value is {type(row).__name__}, expected an object"
            if kind == "expected":
                die(msg)
            print(f"score.py: {msg} — dropped", file=sys.stderr)
            continue
        rid = row.get("request_id")
        if not hashable(rid):
            msg = f"{path} line {n}: request_id is not a scalar ({rid!r})"
            if kind == "expected":
                die(msg)
            print(f"score.py: {msg} — dropped", file=sys.stderr)
            continue
        if kind != "expected" and rid is not None and not isinstance(rid, str):
            row = {**row, "request_id": str(rid)}
        out.append(row)
    if out:                                  # cheap shape check: catches swapped arguments
        looks_actual = sum(1 for o in out if {"confidence", "route", "unknowns"} & set(o))
        looks_key = sum(1 for o in out if {"forbidden_kinds", "alternates", "clean",
                                           "negative_control"} & set(o)
                        or any(isinstance(c, dict) and "min_severity" in c
                               for c in as_list(o.get("conflicts"))))
        if kind == "expected" and looks_actual > len(out) / 2 and not looks_key:
            print(f"score.py: WARNING {path} looks like agent output, not a key — "
                  f"did you swap the arguments?", file=sys.stderr)
        if kind == "actual" and looks_key > len(out) / 2 and not looks_actual:
            print(f"score.py: WARNING {path} looks like a key, not agent output — "
                  f"did you swap the arguments?", file=sys.stderr)
    dupes = [i for i, c in Counter(o.get("request_id") for o in out).items() if c > 1]
    if dupes:
        if kind == "expected":
            die(f"{path}: duplicate request_id {dupes}")
        print(f"score.py: {path}: duplicate request_id {dupes} — keeping the first of each",
              file=sys.stderr)
        first, deduped = set(), []
        for o in out:
            rid = o.get("request_id")
            if rid in first:
                continue
            first.add(rid)
            deduped.append(o)
        out = deduped
    if kind == "expected":
        for o in out:
            if "request_id" not in o or "verdict" not in o:
                die(f"{path}: an entry is missing request_id or verdict")
            rid = o["request_id"]
            if not isinstance(rid, str):
                die(f"{path}: request_id {rid!r} is not a string")
            for fld in ("conflicts", "alternates", "forbidden_kinds", "site_map"):
                if fld in o and not isinstance(o[fld], list):
                    die(f"{path} {rid}: {fld} must be a list, got {type(o[fld]).__name__}")
            for a in o.get("alternates", []):
                if not isinstance(a, dict):
                    die(f"{path} {rid}: every alternate must be an object")
            for variant in [o] + [{**o, **a} for a in o.get("alternates", [])]:
                if not isinstance(variant.get("conflicts", []), list):
                    die(f"{path} {rid}: conflicts must be a list")
                if accepted(variant) - VERDICTS:
                    die(f"{path} {rid}: unknown verdict")
                for c in variant.get("conflicts", []):
                    if not isinstance(c, dict):
                        die(f"{path} {rid}: every conflict must be an object")
                    if not hashable(c.get("kind")) or c.get("kind") not in KINDS:
                        die(f"{path} {rid}: unknown kind {c.get('kind')!r}")
                    ms = c.get("min_severity", "major")
                    if not hashable(ms) or ms not in SEVERITIES:
                        die(f"{path} {rid}: bad min_severity {ms!r}")
                for sm in variant.get("site_map", []):
                    if not (isinstance(sm, (list, tuple)) and len(sm) >= 2):
                        die(f"{path} {rid}: site_map entries must be [requested, resolved] pairs")
                if not isinstance(variant.get("forbidden_kinds", []), list):
                    die(f"{path} {rid}: forbidden_kinds must be a list")
                for k in variant.get("forbidden_kinds", []):
                    if not hashable(k) or k not in KINDS:
                        die(f"{path} {rid}: unknown forbidden kind {k!r}")
                seen = Counter(c["kind"] for c in variant.get("conflicts", []))
                dup = [k for k, c in seen.items() if c > 1]
                if dup:
                    die(f"{path} {rid}: conflict kind listed twice ({dup}); "
                        f"a kind can only be demanded once")
                clash = seen.keys() & set(variant.get("forbidden_kinds", []))
                if clash:
                    die(f"{path} {rid}: {sorted(clash)} is both expected and forbidden")
    return out


def validate_key(expected):
    """Refuse to quote a score from a key that asserts nothing."""
    problems, notes = [], []
    for e in expected:
        rid = e["request_id"]
        variants = [e] + [{**e, **a} for a in e.get("alternates", [])]
        acc = set().union(*(accepted(v) for v in variants))
        if e.get("conflicts") and any(not v.get("conflicts") for v in variants):
            problems.append(f"{rid}: an alternate waives every expected conflict")
        if any(set(v.get("forbidden_kinds", [])) < set(e.get("forbidden_kinds", []))
               for v in variants):
            problems.append(f"{rid}: an alternate narrows forbidden_kinds")
        if len(acc) > 2:
            problems.append(f"{rid}: {len(acc)} acceptable verdicts — asserts nothing")
        if e.get("negative_control") and not e.get("conflicts") \
                and not any(a.get("conflicts") for a in e.get("alternates", [])):
            problems.append(f"{rid}: negative_control with no expected conflict anywhere")
        elif e.get("negative_control") and not e.get("conflicts"):
            notes.append(f"{rid}: negative control expects a bare refusal; the refusal is the "
                         f"finding, and the conflict lives on the solve-instead alternate")
        base = accepted(e)
        if e.get("negative_control") and not base <= REFUSALS:
            problems.append(f"{rid}: negative_control whose BASE verdict is not a refusal")
        elif e.get("negative_control") and not acc <= REFUSALS:
            notes.append(f"{rid}: negative control has a non-refusal alternate; the metric "
                         f"scores only the refusal answers")
        for v in variants:
            vacc = accepted(v)
            floors = [c.get("min_severity", "major") for c in v.get("conflicts", [])]
            if "blocking" in floors and not (vacc & REFUSALS):
                problems.append(f"{rid}: a variant demands a blocking conflict but accepts "
                                f"only {sorted(vacc)} — unreachable at 2/2")
            elif ("major" in floors or "blocking" in floors) and vacc <= {"feasible"}:
                problems.append(f"{rid}: a variant demands a major conflict but accepts only "
                                f"'feasible' — the incoherence cap makes it unreachable at 2/2")
    n = len(expected)
    n_conf = sum(1 for e in expected if e.get("conflicts"))
    if n and not any(e.get("forbidden_kinds") for e in expected):
        notes.append("no case forbids any kind — a shotgun answer cannot be caught by points; "
                     "read shotgun_index and clean_case_precision")
    if n and n_conf < max(1, n // 3):
        notes.append(f"only {n_conf}/{n} cases expect any conflict — expected for the "
                     f"supplied dev key, a warning for a self-authored one")
    if n:
        sb = stub_baseline(expected)
        if sb >= 0.75:
            problems.append(f"a constant feasible/no-conflict stub scores {sb} against this "
                            f"key — it does not measure enough to quote")
        elif sb >= 0.34:
            notes.append(f"stub_baseline {sb}: a constant answer already earns that much; "
                         f"read the sub-metrics, not the headline")
    return {"cases": n, "with_conflicts": n_conf, "notes": notes,
            "clean": sum(1 for e in expected if e.get("clean")),
            "negative_controls": sum(1 for e in expected if e.get("negative_control")),
            "kinds": dict(Counter(c["kind"] for e in expected for c in e.get("conflicts", []))),
            "problems": problems}


def _score_one(exp, act, boiler=frozenset(), allowance=None):
    exp_conf = dicts(exp.get("conflicts"))
    floor = {}
    for c in exp_conf:                       # a kind demanded twice is one demand
        k = c["kind"]
        floor[k] = max(floor.get(k, -1), RANK[c.get("min_severity", "major")])
    exp_kinds = Counter(floor.keys())      # each demanded kind counts once
    n_exp = len(floor)

    if act is None:
        return 0, "no output", 0, 0, n_exp, []

    act_conf = clean_conflicts(act)
    best = {}
    for c in act_conf:
        best[c["kind"]] = max(best.get(c["kind"], -1), RANK[sev_of(c)])

    qualifying = Counter(k for k in exp_kinds if best.get(k, -1) >= floor[k])
    matched = exp_kinds & qualifying
    n_match = sum(matched.values())
    unexpected = sorted(set(best) - set(exp_kinds))
    tp, fn, fp = n_match, n_exp - n_match, 0   # extras are reported, never charged as fp
    # Extras are free, but indiscriminate is not an answer. Two beyond what the key
    # demands is generous room to see more than the key; six is a switch left on.
    discriminate_ok = len(best) <= (n_exp + 2 if allowance is None else allowance)

    av = act.get("verdict")
    verdict_ok = hashable(av) and av in accepted(exp)
    forbidden = {k for k in as_list(exp.get("forbidden_kinds")) if hashable(k)}
    hit_forbidden = sorted(set(best) & forbidden)   # forbidden at any severity

    need_res = {c["kind"] for c in exp_conf
                if c.get("resolution_required",
                         RANK[c.get("min_severity", "major")] < 2)}
    def real_resolution(x):
        r = x.get("resolution")
        if not isinstance(r, str):          # a list/int/dict is not a resolution
            return False
        t = flatten_text(r)
        return bool(t) and t not in boiler

    resolved_ok = all(
        any(real_resolution(x) for x in act_conf if x["kind"] == k)
        for k in need_res)

    conflicts_ok = ((n_match == n_exp) and resolved_ok
                    and not hit_forbidden and discriminate_ok)

    # A correct verdict with none of the reasoning behind it is a guess, not an answer:
    # it pays only where the key expects no conflict at all.
    verdict_pays = verdict_ok and (n_exp == 0 or n_match > 0)

    if verdict_ok and conflicts_ok:
        pts, why = 2, "full"
    elif verdict_pays or (conflicts_ok and n_exp > 0):
        pts, why = 1, "partial"
    else:
        pts, why = 0, "miss"

    sev = {sev_of(c) for c in act_conf}
    if (av == "feasible" and (sev & {"blocking", "major"})) or \
       ("blocking" in sev and (not hashable(av) or av not in REFUSALS)):
        pts, why = min(pts, 1), why + "+incoherent"
    if hit_forbidden:
        why += "+forbidden" + str(hit_forbidden)
    if not discriminate_ok:
        why += f"+indiscriminate({len(best)} kinds, key expects {n_exp})"
    return pts, why, tp, fp, fn, unexpected


def score_case(exp, act, boiler=frozenset()):
    cands = [exp] + [{**exp, **alt} for alt in exp.get("alternates", [])]
    # One allowance for the whole case, set by the narrowest variant. Otherwise an
    # alternate that expects more kinds silently widens the gate for every other branch.
    allowance = min(len({c["kind"] for c in dicts(v.get("conflicts"))}) for v in cands) + 2
    # ties broken toward the variant the candidate actually answered, so that
    # `conflict_recall` and `unexpected_kinds` describe his answer, not declaration order
    out = max((_score_one(e, act, boiler, allowance) for e in cands),
              key=lambda r: (r[0], r[2], -r[4]))
    # A negative control is passed by refusing. Solving it instead may well be right,
    # but no scorer can check that the alternative route was named properly, so a
    # non-refusal answer is capped at 1 and sent to the hand-read.
    if exp.get("negative_control"):
        av = act.get("verdict") if isinstance(act, dict) else None
        if not (hashable(av) and av in REFUSALS) and out[0] > 1:
            out = (1, out[1] + "+neg_control_not_refused") + out[2:]
    return out


STUB = {"verdict": "feasible", "confidence": "high", "conflicts": [],
        "resolved_sequence": None, "resolved_annotations": {}, "site_map": [],
        "route": [], "unknowns": []}


def stub_baseline(expected):
    """What a constant feasible/no-conflict output scores against this key.
    If this is high, the key is not a measurement."""
    fake = [{**STUB, "request_id": e["request_id"]} for e in expected]
    return score(expected, fake)["score"]


def score(expected, actual):
    by_id = {a.get("request_id"): a for a in actual
             if isinstance(a, dict) and hashable(a.get("request_id"))}
    boiler = boilerplate(actual, len(expected))
    pts = tp = fp = fn = 0
    per_case, missing = [], []
    unexpected_all = Counter()
    dropped_all = Counter()
    hard_unexpected = total_conflicts = 0
    clean_t = clean_ok = neg_t = neg_ok = 0
    seq_t = seq_ok = map_t = map_ok = map_extra = 0
    conf_hits, conf_tot = Counter(), Counter()

    for exp in expected:
        rid = exp["request_id"]
        act = by_id.get(rid)
        if act is None:
            missing.append(rid)
        p, why, a, b, c, unexp = score_case(exp, act, boiler)
        dropped_all.update(unknown_kinds(act))
        pts += p; tp += a; fp += b; fn += c
        unexpected_all.update(unexp)
        if act is not None:
            hard_unexpected += len(unexp)
            total_conflicts += len(clean_conflicts(act))
        per_case.append({"request_id": rid, "points": p, "reason": why,
                         "unexpected_kinds": unexp})

        conf = clean_conflicts(act) if act else []
        if exp.get("clean"):
            clean_t += 1
            clean_ok += 0 if act is None else int(
                not any(sev_of(x) in ("blocking", "major") for x in conf))
        if exp.get("negative_control"):
            neg_t += 1
            kinds_ok = (not exp.get("conflicts")) or bool(
                {x["kind"] for x in conf} & {c["kind"] for c in dicts(exp["conflicts"])})
            av = act.get("verdict") if act else None
            neg_ok += 0 if act is None else int(
                hashable(av) and av in (accepted(exp) & REFUSALS) and kinds_ok)
        if exp.get("resolved_sequence") is not None:
            seq_t += 1
            seq_ok += 0 if act is None else int(
                "resolved_sequence" in act
                and norm(act.get("resolved_sequence")) == norm(exp["resolved_sequence"]))
        if exp.get("site_map"):          # a key that lists no pairs asserts nothing here
            map_t += 1
            got = {(norm(m.get("requested")), norm(m.get("resolved")))
                   for m in dicts(act.get("site_map"))} if act else set()
            want = {(norm(x[0]), norm(x[1]))
                    for x in as_list(exp["site_map"])
                    if isinstance(x, (list, tuple)) and len(x) >= 2}
            # A few extra correct pairs never hurt; spraying every possible pair
            # is not an answer, so the extras are bounded.
            map_ok += int(act is not None and want <= got
                          and len(got) <= len(want) + 3)
            map_extra += len(got - want)
        if act is not None:
            cf = act.get("confidence") or "unset"
            if not hashable(cf):
                cf = "unset"
            cf = str(cf)
            conf_tot[cf] += 1
            conf_hits[cf] += int(p == 2)

    n = len(expected)
    return {
        "cases": n, "points": pts, "max_points": 2 * n,
        "score": round(pts / (2 * n), 4) if n else 0.0,
        "conflict_recall": round(tp / (tp + fn), 4) if (tp + fn) else None,
        "clean_case_precision": round(clean_ok / clean_t, 4) if clean_t else None,
        "negative_control_recall": round(neg_ok / neg_t, 4) if neg_t else None,
        "resolved_sequence_exact": round(seq_ok / seq_t, 4) if seq_t else None,
        "site_map_exact": round(map_ok / map_t, 4) if map_t else None,
        "site_map_extra_pairs": map_extra,
        "accuracy_by_confidence": {k: round(conf_hits[k] / v, 4) for k, v in conf_tot.items()},
        "unexpected_kinds_raised": dict(unexpected_all),
        "unreadable_conflict_kinds": dict(dropped_all),
        "boilerplate_strings": len(boiler),
        "shotgun_index": round(hard_unexpected / n, 3) if n else None,
        "conflicts_per_case": round(total_conflicts / n, 2) if n else None,
        "missing_outputs": missing,
        "unexpected_outputs": sorted((set(by_id) - {e["request_id"] for e in expected}),
                                     key=lambda x: (x is None, str(x))),
        "per_case": per_case,
    }


def brier(forecasts, results):
    if not isinstance(results, dict):
        die(f"results.json: expected an object of request_id -> 0/1/2, got "
            f"{type(results).__name__} (generate it with `score.py --results`)")
    if not isinstance(forecasts, dict):
        die(f"forecasts.json: expected an object of request_id -> probability, got "
            f"{type(forecasts).__name__}")
    for k, v in results.items():
        if isinstance(v, bool) or not isinstance(v, int) or v not in (0, 1, 2):
            die(f"results.json: {k} = {v!r}; expected integer 0/1/2 "
                f"(generate it with `score.py --results`)")
    ids = sorted(results)
    if not ids:
        return {"n": 0, "brier": None}
    imputed, coerced, p = [], [], {}
    for i in ids:
        v = forecasts.get(i)
        if v is None:
            imputed.append(i); p[i] = 0.5; continue
        if isinstance(v, dict):
            v = v.get("p")
        if isinstance(v, bool):                  # true/false is not a probability
            coerced.append(i); p[i] = 0.5; continue
        try:
            f = float(v)
            if f != f or f in (float("inf"), float("-inf")):
                raise ValueError("not finite")
            if not 0.0 <= f <= 1.0:              # percentages, or a sign slip
                coerced.append(i); p[i] = min(1.0, max(0.0, f)); continue
            p[i] = f
        except (TypeError, ValueError):
            coerced.append(i); p[i] = 0.5
    y = {i: 1.0 if results[i] == 2 else 0.0 for i in ids}
    b = sum((p[i] - y[i]) ** 2 for i in ids) / len(ids)
    mean_p = sum(p.values()) / len(ids)
    flat = sum((mean_p - y[i]) ** 2 for i in ids) / len(ids)
    spread = (sum((v - mean_p) ** 2 for v in p.values()) / len(ids)) ** 0.5
    ybar = sum(y.values()) / len(ids)
    best_const = ybar * (1 - ybar)      # what the best possible CONSTANT forecast scores
    out = {"n": len(ids), "brier": round(b, 4),
           "gain_vs_own_mean": round(flat - b, 4),          # <- the honest headline
           "baseline_best_constant": round(best_const, 4),
           "gain_vs_best_constant": round(best_const - b, 4),
           "baseline_uninformed_0.5": 0.25,
           "baseline_his_own_mean": round(flat, 4),
           "discrimination_gain": round(0.25 - b, 4),
           "forecast_spread": round(spread, 4), "outcome_base_rate": round(ybar, 4),
           "imputed_missing_forecasts": imputed, "coerced_forecasts": coerced,
           "note": "READ gain_vs_own_mean AND gain_vs_best_constant, not discrimination_gain: "
                   "when the pass rate is far from 0.5 a flat confident forecast beats the "
                   "fixed 0.25 baseline while predicting nothing. Only per-case discrimination "
                   "moves the two gain_vs_ numbers. forecast_spread < 0.05 is a flat forecast. "
                   "n is small."}
    if len(imputed) + len(coerced) > len(ids) / 3:
        out["INVALID"] = "more than a third of forecasts missing or unusable"
    return out


def agreement(expected_ids, runs):
    per = {}
    for rid in expected_ids:
        vs = [(r[rid].get("verdict") if hashable(r[rid].get("verdict"))
               else "<unhashable>") if rid in r else "<missing>" for r in runs]
        sets = [{c["kind"] for c in clean_conflicts(r[rid])} if rid in r
                else {"<missing>"} for r in runs]
        va = max(Counter(vs).values()) / len(vs)
        js = []
        for i in range(len(sets)):
            for j in range(i + 1, len(sets)):
                u = sets[i] | sets[j]
                if not u:
                    js.append(1.0)
                elif "<missing>" in u:
                    js.append(0.0)
                else:
                    js.append(len(sets[i] & sets[j]) / len(u))
        per[rid] = {"verdict_agreement": round(va, 3),
                    "kind_stability": round(sum(js) / len(js), 3) if js else 1.0,
                    "verdicts": vs}
    return {"per_case": per,
            "stability": round(sum(v["verdict_agreement"] * v["kind_stability"]
                                   for v in per.values()) / len(per), 4) if per else None}


def audit(actual, workbook):
    try:
        import openpyxl
    except ImportError:
        die("openpyxl required for --audit")
    books = {}
    for wbp in (workbook if isinstance(workbook, list) else [workbook]):
        stem = Path(wbp).stem
        if stem in books:
            die(f"two workbooks share the stem {stem!r}; refs would be ambiguous")
        try:
            books[stem] = openpyxl.load_workbook(wbp, data_only=True)
        except OSError as e:
            die(str(e))
        except Exception as e:
            die(f"{wbp}: not a readable .xlsx workbook ({type(e).__name__}: {e})")
    # Only genuine column-header rows. Title and subtitle rows carry the claims a
    # candidate must cite when reporting an erratum, so they are content.
    HEADER_ROWS = {"AA_Index": {4}, "Reaction_Master": {4}, "Order_of_Operations": {4},
                   "Caveats_From_Deck": {4}, "Reagent_Checklist": {4}}
    # RISKS / ALTERNATIVES / CONDITIONS rows carry real content after their label
    # and are the corpus's most valuable rows — they are NOT headers.
    HEAD = ("REAGENTS NAMED", "REACTION STEPS", "REVIEWER NOTES", "STEP |", "# |")
    res = {"resolved": 0, "inference_refs_checked": 0, "unsourced_blocks": 0, "uncited_inference_blocks": 0,
           "dead_refs": [], "header_row_refs": [], "empty_row_refs": [],
           "wrong_workbook_refs": []}

    def check(ref):
        parts = str(ref).split(":")
        if len(parts) != 3 or not parts[2].isascii() or not parts[2].isdigit():
            res["dead_refs"].append(ref); return
        wbname, sheet, row = parts[0], parts[1], int(parts[2])
        if wbname not in books:
            res["wrong_workbook_refs"].append(ref); return
        wb = books[wbname]
        if sheet not in wb.sheetnames:
            res["dead_refs"].append(ref); return
        ws = wb[sheet]
        if not (1 <= row <= ws.max_row):
            res["dead_refs"].append(ref); return
        text = " | ".join(str(c.value) for c in ws[row] if c.value is not None)
        if not text.strip():
            res["empty_row_refs"].append(ref)
        elif row in HEADER_ROWS.get(sheet, set()) or \
                any(text.upper().lstrip().startswith(h) for h in HEAD):
            res["header_row_refs"].append(ref)
        else:
            res["resolved"] += 1

    for obj in actual:
        for block in dicts(obj.get("route")) + dicts(obj.get("conflicts")):
            pv = block.get("provenance")
            pv = pv if isinstance(pv, list) else ([pv] if pv else [])
            if not pv:
                res["unsourced_blocks"] += 1
                continue
            for p in pv:
                if not isinstance(p, dict):
                    res["dead_refs"].append(str(p)); continue
                if p.get("kind") == "corpus":
                    check(p.get("ref", ""))
                elif p.get("kind") == "inference":
                    refs = as_list(p.get("refs"))
                    if not refs:
                        res["uncited_inference_blocks"] += 1
                    for r in refs:
                        res["inference_refs_checked"] += 1
                        check(r)
    return res


def applicability(requests, actual):
    """Flag conflicts whose text names chemistry the request does not contain.

    This is the failure a rules engine actually produces — a rule firing on the
    wrong request, cited to a real row that supports the sentence but not the
    case. --audit cannot see it: it checks that a citation is reachable, a human
    checks that it supports the claim, and nothing else checks that the claim
    applies. Findings here are for a human to read, not points.
    """
    MARKERS = {
        "mtt": "lipidation", "palmito": "lipidation", "diacid": "lipidation",
        "fatty ac": "lipidation",
        "grubbs": "hydrocarbon_stapling", "metathes": "hydrocarbon_stapling",
        "staple": "hydrocarbon_stapling",
        "pegylat": "pegylation", "mpeg": "pegylation", "glycos": "glycosylation", "sugar": "glycosylation",
        "disulfid": "disulfide", "acm": "disulfide", "iodine": "disulfide",
        "lactam": "cyclization", "macrocyc": "cyclization", "cycliz": "cyclization", "cyclis": "cyclization",
        "retro-inverso": "retro_inverso", "n-methyl": "n_methylation",
    }
    by_req = {r["request_id"]: r for r in requests
              if isinstance(r, dict) and hashable(r.get("request_id"))
              and r.get("request_id") is not None}
    out = []
    for obj in actual:
        req = by_req.get(obj.get("request_id"))
        if not req:
            continue
        fams = {m.get("family") for m in dicts(req.get("modifications"))
                if hashable(m.get("family"))}
        fams |= {w for f in as_list(req.get("parent_features")) for w in
                 (["cyclization"] if "lactam" in str(f).lower() else
                  ["disulfide"] if "disulfid" in str(f).lower() else [])}
        for c in dicts(obj.get("conflicts")):
            # Only the description. A resolution legitimately names chemistry the
            # request does not contain — that is what a resolution is for.
            # An intent_not_achieved description has to name the alternative
            # mechanism to be an argument at all — exempt it.
            if c.get("kind") == "intent_not_achieved":
                continue
            text = str(c.get('description', '')).lower()
            for marker, fam in MARKERS.items():
                if marker in text and fam not in fams:
                    out.append({"request_id": obj["request_id"], "kind": c.get("kind"),
                                "severity": c.get("severity"),
                                "mentions": marker, "implies_family": fam,
                                "request_families": sorted(str(f) for f in fams if f)})
                    break
    return {"suspect_conflicts": out, "count": len(out)}


def main():
    try:
        signal.signal(signal.SIGPIPE, signal.SIG_DFL)
    except (AttributeError, ValueError):
        pass
    a = sys.argv[1:]

    def jload(path):
        try:
            with open(path, encoding="utf-8") as fh:
                return json.load(fh)
        except UnicodeDecodeError as e:
            die(f"{path}: not valid UTF-8 ({e})")
        except json.JSONDecodeError as e:
            die(f"{path}: not valid JSON ({e})")
        except OSError as e:
            die(str(e))

    if a and a[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)
    KNOWN = {"--json", "--validate", "--validate-key", "--results", "--brier",
             "--agreement", "--audit", "--applies"}
    for x in a:
        if x.startswith("-") and x not in KNOWN:
            die(f"unknown option {x!r}")
    try:
        if a and a[0] == "--brier":
            if len(a) < 3: die("--brier needs forecasts.json results.json")
            if len(a) > 3: die("--brier takes exactly two files")
            print(json.dumps(brier(jload(a[1]), jload(a[2])), indent=2))
        elif a and a[0] == "--agreement":
            if len(a) < 4: die("--agreement needs expected.jsonl and at least two runs")
            ids = [e["request_id"] for e in load(a[1], "expected")]
            runs = [{o["request_id"]: o for o in load(p, "actual")
                     if o.get("request_id") is not None} for p in a[2:]]
            print(json.dumps(agreement(ids, runs), indent=2))
        elif a and a[0] == "--applies":
            if len(a) < 3: die("--applies needs requests.jsonl actual.jsonl")
            if len(a) > 3: die("--applies takes exactly two files")
            out = applicability(load(a[1], "actual"), load(a[2], "actual"))
            print(json.dumps(out, indent=2))
            # advisory only: a comparative reference is not a misapplied conflict.
            # Read the hits; do not gate on them.
        elif a and a[0] == "--audit":
            if len(a) < 3: die("--audit needs actual.jsonl and one or more workbooks")
            out = audit(load(a[1], "actual"), a[2:])
            print(json.dumps(out, indent=2))
            total = out["resolved"] + out["uncited_inference_blocks"] + out["unsourced_blocks"]
            sys.exit(1 if (out["dead_refs"] or out["wrong_workbook_refs"]
                           or out["header_row_refs"] or out["empty_row_refs"]
                           or (total and (out["uncited_inference_blocks"]
                                          + out["unsourced_blocks"]) / total > 0.5)) else 0)
        elif a and a[0] == "--validate-key":
            if len(a) < 2: die("--validate-key needs expected.jsonl")
            if len(a) > 2: die("--validate-key takes exactly one file")
            out = validate_key(load(a[1], "expected"))
            print(json.dumps(out, indent=2))
            sys.exit(1 if out["problems"] else 0)
        elif a and a[0] == "--validate":
            if len(a) < 3: die("--validate needs actual.jsonl schema.json")
            if len(a) > 3: die("--validate takes exactly two files")
            try:
                import jsonschema
            except ImportError:
                die("jsonschema required for --validate (pip install jsonschema)")
            sch = jload(a[2])
            if not isinstance(sch, dict):
                die(f"{a[2]}: schema must be a JSON object")
            bad = []
            try:
                jsonschema.Draft202012Validator.check_schema(sch)
            except Exception as e:
                die(f"{a[2]}: invalid JSON Schema ({e})")
            v = jsonschema.Draft202012Validator(sch)
            with open(a[1], encoding="utf-8-sig") as fh:     # RAW: load() repairs, validate must not
                for n, line in enumerate(fh, 1):
                    if not line.strip():
                        continue
                    try:
                        o = json.loads(line)
                    except json.JSONDecodeError as e:
                        bad.append({"line": n, "error": f"unparseable JSON ({e})"})
                        continue
                    for e in sorted(v.iter_errors(o), key=str):   # every error, not the first
                        bad.append({"line": n,
                                    "request_id": o.get("request_id") if isinstance(o, dict) else None,
                                    "error": e.message})
            print(json.dumps({"checked": True, "invalid": bad}, indent=2))
            sys.exit(1 if bad else 0)
        elif a and a[0] == "--results":
            if len(a) < 3: die("--results needs expected.jsonl actual.jsonl")
            if len(a) > 3: die("--results takes exactly two files")
            s = score(load(a[1], "expected"), load(a[2], "actual"))
            print(json.dumps({c["request_id"]: c["points"] for c in s["per_case"]}, indent=2))
        elif len(a) >= 2:
            as_json = "--json" in a
            a = [x for x in a if x != "--json"]
            if len(a) != 2:
                die("usage: score.py expected.jsonl actual.jsonl [--json]")
            exp_rows = load(a[0], "expected")
            out = score(exp_rows, load(a[1], "actual"))
            out["stub_baseline"] = stub_baseline(exp_rows)
            if as_json:
                print(json.dumps(out, indent=2))
            else:
                pc = out.pop("per_case")
                for k, v in out.items():
                    print(f"{k:26} {v}")
                print("\nper case:")
                for c in pc:
                    extra = f"  +{c['unexpected_kinds']}" if c["unexpected_kinds"] else ""
                    print(f"  {c['request_id']:10} {c['points']}  {c['reason']}{extra}")
        else:
            print(__doc__); sys.exit(1)
    except OSError as e:
        die(str(e))
    except UnicodeDecodeError as e:
        die(f"input is not valid UTF-8: {e}")


if __name__ == "__main__":
    main()
