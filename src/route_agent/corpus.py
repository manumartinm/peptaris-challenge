from __future__ import annotations

import json
from pathlib import Path

from openpyxl import load_workbook  # type: ignore[import-untyped]

from route_agent.models.agent import CitedFact, ProcessProfile
from route_agent.models.corpus import (
    CorpusExcerpt,
    CorpusExcerptRef,
    ExtractedFamiliesView,
    FamilyBinding,
    FamilyProfileView,
    Provenance,
    TargetLookupResult,
)
from route_agent.models.request import DesignRequest
from route_agent.models.validation import (
    ErrorCode,
    ValidationCheck,
    ValidationError,
    ValidationStage,
)
from route_agent.parser.errors import ErrorFactory

SUPPORTED_FAMILIES_SCHEMA = "2.0.0"


def collect_excerpt_refs(node: object) -> list[CorpusExcerptRef]:
    found: list[CorpusExcerptRef] = []
    if isinstance(node, dict):
        provenance = node.get("provenance")
        if isinstance(provenance, list):
            for item in provenance:
                if isinstance(item, dict) and item.get("ref"):
                    found.append(
                        CorpusExcerptRef(
                            ref=str(item["ref"]),
                            ref_row=node.get("ref_row"),
                            source_excerpt=str(node.get("source_excerpt") or ""),
                        )
                    )
        for value in node.values():
            found.extend(collect_excerpt_refs(value))
    elif isinstance(node, list):
        for item in node:
            found.extend(collect_excerpt_refs(item))
    return found


class CorpusRepository:
    def __init__(
        self,
        path: Path,
        errors: ErrorFactory | None = None,
        targets_path: Path | None = None,
    ) -> None:
        self._path = path
        self._errors = errors or ErrorFactory()
        self._targets_path = targets_path
        self._data: ExtractedFamiliesView | None = None
        self._profiles: dict[tuple[str, str], ProcessProfile] = {}
        self._target_rows: dict[str, dict[str, str | None]] | None = None

    def _load_family_catalog(self) -> ExtractedFamiliesView:
        if self._data is None:
            payload = json.loads(self._path.read_text(encoding="utf-8"))
            version = str(payload.get("schema_version") or "")
            if version != SUPPORTED_FAMILIES_SCHEMA:
                raise ValueError(
                    f"unsupported extracted_families schema_version {version!r}; "
                    f"expected {SUPPORTED_FAMILIES_SCHEMA!r}"
                )
            self._data = ExtractedFamiliesView.model_validate(payload)
        return self._data

    def bind_families(
        self, request: DesignRequest
    ) -> tuple[tuple[FamilyBinding, ...], tuple[ValidationError, ...]]:
        catalog = self._load_family_catalog()
        bindings: list[FamilyBinding] = []
        errors: list[ValidationError] = []
        for index, modification in enumerate(request.modifications):
            profile = catalog.families.get(modification.family.value)
            if profile is None:
                errors.append(
                    self._errors.build_error(
                        code=ErrorCode.FAMILY_UNBOUND,
                        check=ValidationCheck.RESOLVE_FAMILY,
                        stage=ValidationStage.RESOLVE_FAMILY,
                        field_path=f"modifications[{index}].family",
                        input_snapshot={
                            "family": modification.family.value,
                            "path": str(self._path),
                        },
                        expected="family profile in extracted_families.json",
                        got=modification.family.value,
                        message=(
                            f"Family {modification.family.value} is not bound in "
                            f"{self._path.name}."
                        ),
                        cause_type="family_unbound",
                        modification_ref=index,
                    )
                )
                continue
            provenance = (
                Provenance(
                    kind="corpus",
                    ref=self._provenance_ref(catalog.source_workbook, profile),
                ),
            )
            process_ids = self._prefer_requested_handles(
                tuple(profile.processes), modification.detail
            )
            bindings.append(
                FamilyBinding(
                    modification_ref=index,
                    family=modification.family,
                    sheet=profile.sheet,
                    process_ids=process_ids,
                    provenance=provenance,
                )
            )
        order_index = {
            name: position for position, name in enumerate(catalog.family_order)
        }
        bindings.sort(
            key=lambda binding: (
                order_index[binding.family.value],
                binding.modification_ref,
            )
        )
        return tuple(bindings), tuple(errors)

    def _prefer_requested_handles(
        self, process_ids: tuple[str, ...], detail: str | None
    ) -> tuple[str, ...]:
        if not detail or not process_ids:
            return process_ids
        haystack = detail.lower().replace("(", "").replace(")", "").replace("-", "")
        ranked: list[tuple[int, int, str]] = []
        for index, process_id in enumerate(process_ids):
            handle = process_id.split("_", 1)[0].lower().replace("-", "")
            preferred = -1 if handle and handle in haystack else 0
            ranked.append((preferred, index, process_id))
        ranked.sort()
        return tuple(process_id for _preferred, _index, process_id in ranked)

    def lookup_family_process(self, family: str, process_id: str) -> ProcessProfile:
        cached = self._profiles.get((family, process_id))
        if cached is not None:
            return cached
        catalog = self._load_family_catalog()
        profile = catalog.families.get(family)
        process = None if profile is None else profile.processes.get(process_id)
        if profile is None or process is None:
            missing = ProcessProfile(family=family, process_id=process_id, found=False)
            self._profiles[(family, process_id)] = missing
            return missing
        labeled = (
            *process.caveats,
            *process.constraints,
            *process.conditions,
        )
        required: list[str] = []
        for excerpt in process.requires:
            fact = self._cited_fact_from_excerpt(excerpt)
            if fact.text:
                required.append(fact.text)
        risks: list[CitedFact] = []
        alternatives: list[CitedFact] = []
        for excerpt in labeled:
            label = self._excerpt_label(excerpt)
            fact = self._cited_fact_from_excerpt(excerpt)
            if label.startswith("RISK"):
                risks.append(fact)
            elif label.startswith("ALTERNATIVE"):
                alternatives.append(fact)
        built = ProcessProfile(
            family=family,
            process_id=process.process_id or process_id,
            found=True,
            name=process.name,
            summary=profile.summary,
            requires=tuple(required),
            reagents=tuple(
                self._cited_fact_from_excerpt(excerpt) for excerpt in process.reagents
            ),
            conditions=tuple(
                self._cited_fact_from_excerpt(excerpt) for excerpt in process.conditions
            ),
            constraints=tuple(
                self._cited_fact_from_excerpt(excerpt)
                for excerpt in process.constraints
            ),
            explicit_risks=tuple(risks),
            explicit_alternatives=tuple(alternatives),
            stage_hint=process.stage_hint,
            building_blocks=process.building_blocks,
        )
        self._profiles[(family, process_id)] = built
        return built

    def lookup_target(self, peptide: str) -> TargetLookupResult:
        rows = self._load_target_rows()
        if rows is None:
            return TargetLookupResult.unavailable(peptide)
        if not rows:
            return TargetLookupResult(
                available=False, peptide=peptide, reason="workbook_empty"
            )
        record = rows.get(peptide.strip().lower())
        if record is None:
            return TargetLookupResult(
                available=False, peptide=peptide, reason="peptide_not_found"
            )
        return TargetLookupResult(
            available=True,
            peptide=record.get("peptide") or peptide,
            receptor_target=record.get("receptor_target"),
            receptor_class=record.get("receptor_class"),
            ligand_role=record.get("ligand_role"),
            invariant_windows=_split_list(record.get("invariant_windows")),
            sar_precedents=_split_list(record.get("sar_precedents")),
        )

    def _load_target_rows(self) -> dict[str, dict[str, str | None]] | None:
        if self._target_rows is not None:
            return self._target_rows
        if self._targets_path is None or not self._targets_path.is_file():
            return None
        workbook = load_workbook(self._targets_path, read_only=True, data_only=True)
        try:
            loaded = _read_target_sheet(workbook[workbook.sheetnames[0]])
            if (
                "Target_Peptide_Master" in workbook.sheetnames
                and workbook.sheetnames[0] != "Target_Peptide_Master"
            ):
                master = _read_target_sheet(workbook["Target_Peptide_Master"])
                for key, row in master.items():
                    if key in loaded:
                        for field in ("invariant_windows", "sar_precedents"):
                            if row.get(field):
                                loaded[key][field] = row[field]
                    else:
                        loaded[key] = row
            self._target_rows = loaded
            return self._target_rows
        finally:
            workbook.close()

    def _provenance_ref(self, source_workbook: str, profile: FamilyProfileView) -> str:
        sheet = profile.sheet
        for process in profile.processes.values():
            for excerpt in (*process.requires, *process.reagents):
                if excerpt.provenance and excerpt.provenance[0].ref:
                    return excerpt.provenance[0].ref
                if excerpt.ref_row is not None:
                    return f"{source_workbook}:{sheet}:{excerpt.ref_row}"
        return f"{source_workbook}:{sheet}:1"

    def _excerpt_label(self, excerpt: CorpusExcerpt) -> str:
        for cell in excerpt.cells:
            if cell.column == "Step" and cell.value:
                return cell.value.upper()
        return (excerpt.source_excerpt or "").split("|", 1)[0].strip().upper()

    def _cited_fact_from_excerpt(self, excerpt: CorpusExcerpt) -> CitedFact:
        text = excerpt.source_excerpt or ""
        if not text:
            values = [cell.value for cell in excerpt.cells if cell.value]
            text = " | ".join(values)
        ref = excerpt.provenance[0].ref if excerpt.provenance else None
        return CitedFact(text=text, ref_row=excerpt.ref_row, ref=ref)


def _read_target_sheet(sheet: object) -> dict[str, dict[str, str | None]]:
    rows = sheet.iter_rows(values_only=True)  # type: ignore[attr-defined]
    header = next(rows, None)
    if header is None:
        return {}
    names = [str(cell or "").strip() for cell in header]
    loaded: dict[str, dict[str, str | None]] = {}
    for row in rows:
        record = {
            names[index]: row[index] if index < len(row) else None
            for index in range(len(names))
        }
        name = str(record.get("Peptide") or "").strip()
        if not name:
            continue
        loaded[name.lower()] = {
            "peptide": name,
            "receptor_target": _as_optional_text(record.get("Receptor Target")),
            "receptor_class": _as_optional_text(record.get("Receptor Class")),
            "ligand_role": _as_optional_text(record.get("Ligand Role")),
            "invariant_windows": _as_optional_text(
                _first_present(record, ("Invariant Windows", "invariant_windows"))
            ),
            "sar_precedents": _as_optional_text(
                _first_present(record, ("SAR Precedents", "SAR", "sar_precedents"))
            ),
        }
    return loaded


def _as_optional_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _first_present(record: dict[str, object], names: tuple[str, ...]) -> object:
    for name in names:
        if name in record and record[name] not in {None, ""}:
            return record[name]
    return None


def _split_list(value: str | None) -> tuple[str, ...]:
    if not value:
        return ()
    parts: list[str] = []
    for chunk in value.replace("\n", ";").split(";"):
        item = chunk.strip()
        if item:
            parts.append(item)
    return tuple(parts)
